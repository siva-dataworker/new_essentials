"""
Excel Export APIs for Construction Management System
Generates Excel reports for labour, materials, budget, and bills
"""
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from .authentication import JWTAuthentication
from .database import fetch_all, fetch_one
from datetime import datetime
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def style_header(ws, row=1):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_adjust_columns(ws):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def export_labour_entries(request, site_id):
    """
    Export labour entries to Excel
    GET /api/export/labour-entries/{site_id}/
    """
    if not OPENPYXL_AVAILABLE:
        return Response({'error': 'Excel export not available. Install openpyxl.'}, 
                       status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    try:
        # Get site info
        site = fetch_one("SELECT site_name, customer_name FROM sites WHERE id = %s", (site_id,))
        if not site:
            return Response({'error': 'Site not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get labour entries
        entries = fetch_all("""
            SELECT 
                le.entry_date,
                le.entry_time,
                le.day_of_week,
                le.labour_type,
                le.labour_count,
                le.notes,
                le.extra_cost,
                le.extra_cost_notes,
                le.is_modified,
                le.modification_reason,
                u.full_name as supervisor_name,
                le.submitted_by_role
            FROM labour_entries le
            JOIN users u ON le.supervisor_id = u.id
            WHERE le.site_id = %s
            ORDER BY le.entry_date DESC, le.entry_time DESC
        """, (site_id,))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Labour Entries"
        
        # Add title
        ws.merge_cells('A1:L1')
        ws['A1'] = f"Labour Entries Report - {site['customer_name']} {site['site_name']}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Add export info
        ws.merge_cells('A2:L2')
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['Date', 'Time', 'Day', 'Labour Type', 'Count', 'Supervisor', 'Role', 
                  'Notes', 'Extra Cost', 'Extra Cost Notes', 'Modified', 'Modification Reason']
        ws.append([])  # Empty row
        ws.append(headers)
        style_header(ws, row=4)
        
        # Data rows
        for entry in entries:
            ws.append([
                str(entry['entry_date']) if entry['entry_date'] else '',
                str(entry['entry_time'])[:8] if entry['entry_time'] else '',
                entry['day_of_week'] or '',
                entry['labour_type'] or '',
                entry['labour_count'] or 0,
                entry['supervisor_name'] or '',
                entry['submitted_by_role'] or '',
                entry['notes'] or '',
                float(entry['extra_cost']) if entry['extra_cost'] else 0,
                entry['extra_cost_notes'] or '',
                'Yes' if entry['is_modified'] else 'No',
                entry['modification_reason'] or '',
            ])
        
        # Add summary
        total_count = sum(e['labour_count'] or 0 for e in entries)
        total_extra_cost = sum(float(e['extra_cost']) if e['extra_cost'] else 0 for e in entries)
        
        ws.append([])
        ws.append(['SUMMARY', '', '', '', total_count, '', '', '', total_extra_cost, '', '', ''])
        
        # Style summary row
        summary_row = ws.max_row
        for cell in ws[summary_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Auto-adjust columns
        auto_adjust_columns(ws)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Labour_Entries_{site['site_name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def export_material_entries(request, site_id):
    """
    Export material entries to Excel
    GET /api/export/material-entries/{site_id}/
    """
    if not OPENPYXL_AVAILABLE:
        return Response({'error': 'Excel export not available. Install openpyxl.'}, 
                       status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    try:
        # Get site info
        site = fetch_one("SELECT site_name, customer_name FROM sites WHERE id = %s", (site_id,))
        if not site:
            return Response({'error': 'Site not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get material entries
        entries = fetch_all("""
            SELECT 
                mb.entry_date,
                mb.updated_at,
                mb.day_of_week,
                mb.material_type,
                mb.quantity,
                mb.unit,
                mb.extra_cost,
                mb.extra_cost_notes,
                u.full_name as supervisor_name
            FROM material_balances mb
            JOIN users u ON mb.supervisor_id = u.id
            WHERE mb.site_id = %s
            ORDER BY mb.entry_date DESC, mb.updated_at DESC
        """, (site_id,))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Material Entries"
        
        # Add title
        ws.merge_cells('A1:I1')
        ws['A1'] = f"Material Entries Report - {site['customer_name']} {site['site_name']}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Add export info
        ws.merge_cells('A2:I2')
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['Date', 'Time', 'Day', 'Material Type', 'Quantity', 'Unit', 
                  'Supervisor', 'Extra Cost', 'Extra Cost Notes']
        ws.append([])
        ws.append(headers)
        style_header(ws, row=4)
        
        # Data rows
        for entry in entries:
            ws.append([
                str(entry['entry_date']) if entry['entry_date'] else '',
                str(entry['updated_at'])[:19] if entry['updated_at'] else '',
                entry['day_of_week'] or '',
                entry['material_type'] or '',
                float(entry['quantity']) if entry['quantity'] else 0,
                entry['unit'] or '',
                entry['supervisor_name'] or '',
                float(entry['extra_cost']) if entry['extra_cost'] else 0,
                entry['extra_cost_notes'] or '',
            ])
        
        # Add summary by material type
        material_summary = {}
        for entry in entries:
            mat_type = entry['material_type'] or 'Unknown'
            if mat_type not in material_summary:
                material_summary[mat_type] = {'quantity': 0, 'unit': entry['unit'] or ''}
            material_summary[mat_type]['quantity'] += float(entry['quantity']) if entry['quantity'] else 0
        
        ws.append([])
        ws.append(['SUMMARY BY MATERIAL TYPE'])
        ws.append(['Material Type', 'Total Quantity', 'Unit'])
        
        for mat_type, data in material_summary.items():
            ws.append([mat_type, data['quantity'], data['unit']])
        
        # Auto-adjust columns
        auto_adjust_columns(ws)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Material_Entries_{site['site_name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def export_budget_utilization(request, site_id):
    """
    Export budget utilization to Excel
    GET /api/export/budget-utilization/{site_id}/
    """
    if not OPENPYXL_AVAILABLE:
        return Response({'error': 'Excel export not available. Install openpyxl.'}, 
                       status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    try:
        # Get site info
        site = fetch_one("SELECT site_name, customer_name FROM sites WHERE id = %s", (site_id,))
        if not site:
            return Response({'error': 'Site not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get budget summary
        summary = fetch_one("""
            SELECT * FROM budget_utilization_summary WHERE site_id = %s
        """, (site_id,))
        
        if not summary:
            return Response({'error': 'No budget data found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get material breakdown
        material_costs = fetch_all("""
            SELECT 
                material_type,
                SUM(total_cost) as total_cost,
                SUM(quantity) as total_quantity,
                unit
            FROM material_cost_tracking
            WHERE site_id = %s
            GROUP BY material_type, unit
            ORDER BY total_cost DESC
        """, (site_id,))
        
        # Get labour breakdown
        labour_costs = fetch_all("""
            SELECT 
                labour_type,
                SUM(labour_count) as total_count,
                AVG(daily_rate) as avg_rate,
                SUM(total_cost) as total_cost
            FROM labour_cost_calculation
            WHERE site_id = %s
            GROUP BY labour_type
            ORDER BY total_cost DESC
        """, (site_id,))
        
        # Create workbook
        wb = Workbook()
        
        # Summary Sheet
        ws1 = wb.active
        ws1.title = "Budget Summary"
        
        ws1.merge_cells('A1:B1')
        ws1['A1'] = f"Budget Utilization Report - {site['customer_name']} {site['site_name']}"
        ws1['A1'].font = Font(bold=True, size=14)
        
        ws1.append([])
        ws1.append(['Budget Item', 'Amount (₹)'])
        style_header(ws1, row=3)
        
        ws1.append(['Total Budget', float(summary['total_budget']) if summary['total_budget'] else 0])
        ws1.append(['Material Budget', float(summary['material_budget']) if summary['material_budget'] else 0])
        ws1.append(['Labour Budget', float(summary['labour_budget']) if summary['labour_budget'] else 0])
        ws1.append(['Other Budget', float(summary['other_budget']) if summary['other_budget'] else 0])
        ws1.append([])
        ws1.append(['Total Material Cost', float(summary['total_material_cost']) if summary['total_material_cost'] else 0])
        ws1.append(['Total Labour Cost', float(summary['total_labour_cost']) if summary['total_labour_cost'] else 0])
        ws1.append(['Total Vendor Cost', float(summary['total_vendor_cost']) if summary['total_vendor_cost'] else 0])
        ws1.append([])
        ws1.append(['Total Spent', float(summary['total_spent']) if summary['total_spent'] else 0])
        ws1.append(['Remaining Budget', float(summary['remaining_budget']) if summary['remaining_budget'] else 0])
        ws1.append(['Utilization %', f"{float(summary['utilization_percentage']) if summary['utilization_percentage'] else 0:.2f}%"])
        ws1.append(['Status', summary['status'] or 'N/A'])
        
        auto_adjust_columns(ws1)
        
        # Material Breakdown Sheet
        ws2 = wb.create_sheet("Material Breakdown")
        ws2.append(['Material Type', 'Total Quantity', 'Unit', 'Total Cost (₹)'])
        style_header(ws2, row=1)
        
        for m in material_costs:
            ws2.append([
                m['material_type'] or '',
                float(m['total_quantity']) if m['total_quantity'] else 0,
                m['unit'] or '',
                float(m['total_cost']) if m['total_cost'] else 0,
            ])
        
        auto_adjust_columns(ws2)
        
        # Labour Breakdown Sheet
        ws3 = wb.create_sheet("Labour Breakdown")
        ws3.append(['Labour Type', 'Total Count', 'Avg Daily Rate (₹)', 'Total Cost (₹)'])
        style_header(ws3, row=1)
        
        for l in labour_costs:
            ws3.append([
                l['labour_type'] or '',
                int(l['total_count']) if l['total_count'] else 0,
                float(l['avg_rate']) if l['avg_rate'] else 0,
                float(l['total_cost']) if l['total_cost'] else 0,
            ])
        
        auto_adjust_columns(ws3)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Budget_Utilization_{site['site_name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def export_bills(request, site_id):
    """
    Export bills to Excel
    GET /api/export/bills/{site_id}/
    """
    if not OPENPYXL_AVAILABLE:
        return Response({'error': 'Excel export not available. Install openpyxl.'}, 
                       status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    try:
        # Get site info
        site = fetch_one("SELECT site_name, customer_name FROM sites WHERE id = %s", (site_id,))
        if not site:
            return Response({'error': 'Site not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get bills
        bills = fetch_all("""
            SELECT 
                b.bill_date,
                b.bill_number,
                b.material_type,
                b.quantity,
                b.unit,
                b.price_per_unit,
                b.total_amount,
                b.vendor_name,
                b.payment_status,
                b.paid_amount,
                b.payment_date,
                u.full_name as uploaded_by,
                b.created_at
            FROM bills b
            LEFT JOIN users u ON b.uploaded_by = u.id
            WHERE b.site_id = %s
            ORDER BY b.bill_date DESC, b.created_at DESC
        """, (site_id,))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Bills"
        
        # Add title
        ws.merge_cells('A1:M1')
        ws['A1'] = f"Bills Report - {site['customer_name']} {site['site_name']}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Add export info
        ws.merge_cells('A2:M2')
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['Bill Date', 'Bill Number', 'Material Type', 'Quantity', 'Unit', 
                  'Price/Unit', 'Total Amount', 'Vendor', 'Payment Status', 
                  'Paid Amount', 'Payment Date', 'Uploaded By', 'Created At']
        ws.append([])
        ws.append(headers)
        style_header(ws, row=4)
        
        # Data rows
        total_amount = 0
        total_paid = 0
        
        for bill in bills:
            ws.append([
                str(bill['bill_date']) if bill['bill_date'] else '',
                bill['bill_number'] or '',
                bill['material_type'] or '',
                float(bill['quantity']) if bill['quantity'] else 0,
                bill['unit'] or '',
                float(bill['price_per_unit']) if bill['price_per_unit'] else 0,
                float(bill['total_amount']) if bill['total_amount'] else 0,
                bill['vendor_name'] or '',
                bill['payment_status'] or '',
                float(bill['paid_amount']) if bill['paid_amount'] else 0,
                str(bill['payment_date']) if bill['payment_date'] else '',
                bill['uploaded_by'] or '',
                str(bill['created_at'])[:19] if bill['created_at'] else '',
            ])
            
            total_amount += float(bill['total_amount']) if bill['total_amount'] else 0
            total_paid += float(bill['paid_amount']) if bill['paid_amount'] else 0
        
        # Add summary
        ws.append([])
        ws.append(['SUMMARY', '', '', '', '', '', total_amount, '', '', total_paid, '', '', ''])
        
        # Style summary row
        summary_row = ws.max_row
        for cell in ws[summary_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Auto-adjust columns
        auto_adjust_columns(ws)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Bills_{site['site_name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
